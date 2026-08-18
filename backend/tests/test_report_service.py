from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.report_service import ReportService


def test_report_service_resolves_a_real_cjk_font_for_charts() -> None:
    font_path = ReportService._resolve_chart_font_path()

    assert font_path.is_file()
    assert font_path.suffix.lower() in {".ttc", ".ttf", ".otf"}


def test_report_service_renders_pdf_summary_with_cjk_font() -> None:
    analysis = {
        "metrics": {
            "total_rows": 3,
            "sales_amount": {"total": 3600},
        },
        "summary": "华东区域销售表现领先。",
        "anomalies": ["华北完成率偏低。"],
        "recommendations": ["优先优化华北区域客户覆盖。"],
    }

    image = ReportService._build_pdf_summary_image(SimpleNamespace(original_filename="销售数据.csv"), analysis)

    assert image.getvalue().startswith(b"\x89PNG")


def test_report_service_builds_enterprise_overview_rows() -> None:
    metrics = {
        "sales_amount": {"total": 3600, "average": 1200},
        "completion_rate": 84.44,
        "highest_sales_region": {"name": "\u534e\u4e1c", "value": 2200},
        "lowest_sales_region": {"name": "\u534e\u5317", "value": 600},
        "region_ranking": [
            {"name": "\u534e\u4e1c", "value": 2200},
            {"name": "\u534e\u5357", "value": 800},
            {"name": "\u534e\u5317", "value": 600},
        ],
    }

    rows = dict(ReportService._overview_rows(metrics))

    assert rows["\u9500\u552e\u989d\u603b\u8ba1"] == 3600
    assert rows["\u5e73\u5747\u9500\u552e\u989d"] == 1200
    assert rows["\u6700\u5927\u9500\u552e\u533a\u57df"] == "\u534e\u4e1c\uff082200\uff09"
    assert rows["\u6700\u4f4e\u9500\u552e\u533a\u57df"] == "\u534e\u5317\uff08600\uff09"
    assert rows["\u5b8c\u6210\u7387"] == "84.44%"
    assert "1. \u534e\u4e1c\uff1a2200" in rows["\u533a\u57df TOP \u6392\u540d"]


def test_report_service_omits_unavailable_sales_and_lists_skipped_analyses() -> None:
    metrics = {
        "total_rows": 2,
        "sales_amount": None,
        "completion_rate": None,
        "order_count": 2,
        "product_quantity": [{"name": "A", "value": 5}],
        "region_ranking": [],
        "top_regions": [],
        "analysis_plan": [
            {
                "id": "sales_total",
                "name": "销售额分析",
                "supported": False,
                "missing_fields": ["sales_amount"],
                "reason": "缺少所需字段",
            }
        ],
    }

    rows = dict(ReportService._overview_rows(metrics))

    assert "销售额总计" not in rows
    assert rows["订单数量"] == 2
    assert "商品销量排名" in rows
    assert ReportService._skipped_analysis_rows(metrics) == ["销售额分析：缺少 sales_amount 字段"]


def test_report_service_keeps_real_zero_sales_in_overview() -> None:
    metrics = {
        "sales_amount": {"total": 0, "average": 0},
        "completion_rate": None,
        "region_ranking": [],
        "top_regions": [],
        "order_count": None,
        "product_quantity": [],
        "analysis_plan": [],
    }

    rows = dict(ReportService._overview_rows(metrics))

    assert rows["销售额总计"] == 0
    assert rows["平均销售额"] == 0


def student_metrics(**analysis_overrides: object) -> dict:
    score_analysis = {
        "student_count": 2,
        "score_summary": {
            "count": 4,
            "average": 0.0,
            "maximum": 0.0,
            "minimum": 0.0,
            "median": 0.0,
        },
        "subject_score": [
            {"name": "数学", "count": 2, "average": 0.0, "maximum": 0.0, "minimum": 0.0},
        ],
        "class_score": [],
        "student_score": [
            {"student_id": "S-1", "score_count": 2, "average": 0.0, "maximum": 0.0, "minimum": 0.0},
        ],
        "exam_trend": [],
    }
    score_analysis.update(analysis_overrides)
    return {
        "total_rows": 4,
        "selected_module": {"id": "student_score", "name": "学生成绩分析"},
        "student_score_analysis": score_analysis,
        "sales_amount": None,
        "completion_rate": None,
        "order_count": None,
        "product_quantity": [],
        "region_ranking": [],
        "top_regions": [],
        "analysis_plan": [],
    }


def test_student_score_overview_contains_real_score_metrics_without_order_metrics() -> None:
    rows = dict(ReportService._overview_rows(student_metrics()))

    assert rows["学生数量"] == 2
    assert rows["有效成绩数量"] == 4
    assert rows["平均分"] == 0.0
    assert rows["中位数"] == 0.0
    assert rows["最高分"] == 0.0
    assert rows["最低分"] == 0.0
    assert "销售额总计" not in rows
    assert "完成率" not in rows


def test_student_score_exports_create_dynamic_tables_without_order_chart() -> None:
    metrics = student_metrics()
    analysis = {
        "mode": "rule_based",
        "metrics": metrics,
        "summary": "学生数量 2，平均分 0.0。",
        "anomalies": ["已计算的成绩指标未触发额外风险规则。"],
        "business_problems": ["应继续积累考试数据。"],
        "recommendations": ["围绕已存在的学科成绩安排复盘。"],
    }
    service = ReportService()
    service.analysis_service = SimpleNamespace(generate_report=lambda *_args: analysis)
    dataset = SimpleNamespace(original_filename="scores.csv")

    workbook = load_workbook(BytesIO(service.build_excel(None, dataset)))
    values = [cell.value for row in workbook.active.iter_rows() for cell in row if cell.value is not None]

    assert "学生数量" in values
    assert "学科成绩统计" in values
    assert "班级成绩统计" not in values
    assert "销售额总计" not in values
    assert workbook.active._charts == []
    assert service.build_word(None, dataset).startswith(b"PK")
    assert service.build_pdf(None, dataset).startswith(b"%PDF")


def test_report_exports_skip_unavailable_sales_and_region_chart() -> None:
    metrics = {
        "total_rows": 2,
        "sales_amount": None,
        "completion_rate": None,
        "order_count": 2,
        "product_quantity": [{"name": "A", "value": 5}],
        "region_ranking": [],
        "top_regions": [],
        "analysis_plan": [
            {
                "id": "sales_total",
                "name": "销售额分析",
                "supported": False,
                "missing_fields": ["sales_amount"],
                "reason": "缺少所需字段",
            }
        ],
    }
    analysis = {
        "mode": "rule_based",
        "metrics": metrics,
        "summary": "可统计订单与商品销量。",
        "anomalies": ["未触发预设风险规则。"],
        "business_problems": ["需持续积累数据。"],
        "recommendations": ["按周复盘。"],
    }
    service = ReportService()
    service.analysis_service = SimpleNamespace(generate_report=lambda *_args: analysis)
    dataset = SimpleNamespace(original_filename="minimal.csv")

    excel = load_workbook(BytesIO(service.build_excel(None, dataset)))
    labels = [excel.active.cell(row=row, column=1).value for row in range(1, excel.active.max_row + 1)]

    assert "销售额总计" not in labels
    assert excel.active._charts == []
    assert service.build_word(None, dataset).startswith(b"PK")
    assert service.build_pdf(None, dataset).startswith(b"%PDF")


def inventory_metrics() -> dict:
    return {
        "total_rows": 3,
        "selected_module": {"id": "inventory", "name": "库存分析"},
        "inventory_analysis": {
            "inventory_count": 3,
            "stock_summary": {
                "count": 2,
                "total": 35.0,
                "average": 17.5,
                "maximum": 30.0,
                "minimum": 5.0,
                "median": 17.5,
            },
            "low_stock_analysis": [
                {"product_id": "P001", "product_name": "商品A", "stock_quantity": 5.0, "safety_stock": 10.0, "shortage": 5.0}
            ],
            "inventory_value": {"count": 2, "total": 460.0, "average": 230.0},
            "category_stock": [{"name": "电子", "value": 35.0}],
            "warehouse_stock": [{"name": "郑州仓", "value": 35.0}],
            "supplier_stock": [],
            "inventory_flow": None,
            "inventory_trend": [],
        },
        "sales_amount": None,
        "completion_rate": None,
        "order_count": None,
        "product_quantity": [],
        "region_ranking": [],
        "top_regions": [],
        "analysis_plan": [],
    }


def test_inventory_overview_contains_only_real_inventory_metrics() -> None:
    rows = dict(ReportService._overview_rows(inventory_metrics()))

    assert rows["商品数量"] == 3
    assert rows["库存总量"] == 35.0
    assert rows["平均库存"] == 17.5
    assert rows["库存价值总计"] == 460.0
    assert "销售额总计" not in rows
    assert "完成率" not in rows


def test_inventory_exports_render_tables_without_order_chart() -> None:
    metrics = inventory_metrics()
    analysis = {
        "mode": "rule_based",
        "metrics": metrics,
        "summary": "当前库存总量 35.0。",
        "anomalies": ["存在 1 个低库存商品。"],
        "business_problems": ["低库存商品存在库存缺口。"],
        "recommendations": ["核对低库存商品库存记录。"],
    }
    service = ReportService()
    service.analysis_service = SimpleNamespace(generate_report=lambda *_args: analysis)
    dataset = SimpleNamespace(original_filename="inventory.csv")

    workbook = load_workbook(BytesIO(service.build_excel(None, dataset)))
    values = [cell.value for row in workbook.active.iter_rows() for cell in row if cell.value is not None]

    assert "商品数量" in values
    assert "低库存明细" in values
    assert "分类库存统计" in values
    assert "仓库库存统计" in values
    assert "销售额总计" not in values
    assert workbook.active._charts == []
    assert service.build_word(None, dataset).startswith(b"PK")
    assert service.build_pdf(None, dataset).startswith(b"%PDF")


def test_order_exports_use_dynamic_business_sections_and_trusted_amount_notice() -> None:
    metrics = {
        "total_rows": 2,
        "selected_module": {"id": "order", "name": "订单分析"},
        "top_regions": [{"name": "郑州", "value": 100.0}],
        "analysis_plan": [],
        "order_analysis": {
            "overview": {
                "record_count": 2,
                "order_count": 2,
                "sales_total": 100.0,
                "average_order_value": 50.0,
                "verified_sales_total": 100.0,
                "verified_order_count": 2,
                "average_verified_order_value": 50.0,
                "amount_mismatch_count": 1,
            },
            "product_analysis": [{"name": "商品A", "order_count": 2, "quantity": 2.0, "sales_amount": 100.0}],
            "category_analysis": [], "region_analysis": [], "status_analysis": [], "payment_method_analysis": [],
            "customer_analysis": None, "time_analysis": {}, "discount_analysis": None,
            "data_quality": {"amount_mismatch_count": 1},
        },
    }
    analysis = {"mode": "rule_based", "metrics": metrics, "summary": "可信销售额 100。", "anomalies": ["金额不一致。"], "business_problems": ["需复核。"], "recommendations": ["按可信金额统计。"]}
    service = ReportService(); service.analysis_service = SimpleNamespace(generate_report=lambda *_args: analysis)
    workbook = load_workbook(BytesIO(service.build_excel(None, SimpleNamespace(original_filename="orders.xlsx"))))
    values = [cell.value for row in workbook.active.iter_rows() for cell in row if cell.value is not None]

    assert "已验证销售额总计" in values
    assert "统计口径" in values
    assert "商品销售分析" in values
    assert "品类销售分析" not in values
    assert workbook.active._charts
